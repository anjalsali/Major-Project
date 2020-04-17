import cv2
import argparse
import numpy as np
import openpyxl
from openpyxl import load_workbook
from motpy import Detection, MultiObjectTracker
import imutils
import time

def run(videoName,dateVar,timeVar):

	ageProto="../Models/age_deploy.prototxt"
	ageModel="../Models/age_net.caffemodel"
	genderProto="../Models/gender_deploy.prototxt"
	genderModel="../Models/gender_net.caffemodel"
	MODEL_MEAN_VALUES=(78.4263377603, 87.7689143744, 114.895847746)
	ageList=['(0-10)', '(0-10)', '(10-15)', '(15-18)', '(18-25)', '(30-45)', '(48-55)', '(60-100)']
	genderList=['Male','Female']
	ageNet=cv2.dnn.readNet(ageModel,ageProto)
	genderNet=cv2.dnn.readNet(genderModel,genderProto)
	
	net = cv2.dnn.readNetFromCaffe("../Models/deploy.prototxt","../Models/res10_300x300_ssd_iter_140000.caffemodel")
	(H, W) = (None, None)

	detect_interval = 1
	scale_rate = 0.75
	show_rate = 1
	colours = np.random.rand(32, 3)
		
	c=0
	id_dict={}


	wb=openpyxl.load_workbook('./Output/outputgad.xlsx')

	ws=wb.worksheets[0]
	rowno = ws.max_row + 1
	count = ws.max_row - 1

	webcam = cv2.VideoCapture(videoName)
	fps = webcam.get(cv2.CAP_PROP_FPS)

	if fps==0:
		print("No Input Stream Detected")
		webcam.release()
		cv2.destroyAllWindows()
		return 
		
	
	tracker = MultiObjectTracker(
        dt=1/fps, tracker_kwargs={'max_staleness': 3},
        model_spec='constant_acceleration_and_static_box_size_2d',
        matching_fn_kwargs={'min_iou': 0.25})

	if not webcam.isOpened():
		print("No Input Stream Detected")
		webcam.release()
		cv2.destroyAllWindows()
		return

	if(videoName==0):
		frameWidth=400
		padding=25
		threshold = 0.5
	else:
		frameWidth=1000
		padding=20
		threshold = 0.8

	final_faces = []
	while(webcam.isOpened()):
		status, frame = webcam.read()

		if frame is None:
			print("Could not read frame")
			webcam.release()
			cv2.destroyAllWindows()
			break
	

		frame = imutils.resize(frame, width=frameWidth)


		(H, W) = frame.shape[:2]
		#frame = cv2.resize(frame, (0, 0), fx=scale_rate, fy=scale_rate)
		if not status:
			print("Could not read frame")
			webcam.release()
			cv2.destroyAllWindows()
			break

		#gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
		
		blob = cv2.dnn.blobFromImage(frame, 1.0, (W, H),(104.0, 177.0, 123.0))
		net.setInput(blob)
		detections = net.forward()
		rects = []

		for i in range(0, detections.shape[2]):
		# filter out weak detections by ensuring the predicted
		# probability is greater than a minimum threshold
			if detections[0, 0, i, 2] > threshold:
			# compute the (x, y)-coordinates of the bounding box for
			# the object, then update the bounding box rectangles list
				box = detections[0, 0, i, 3:7] * np.array([W, H, W, H])
				rects.append(box.astype("int"))

	
			face_list=[]
			for item in rects:
				xmin = item[0]
				ymin = item[1]
				xmax = item[0] + item[2]
				ymax = item[1] + item[3]
				face_list.append([xmin, ymin, xmax, ymax])
	
			final_faces = np.array(face_list) 

		
		
		detections = [Detection(box=bbox) for bbox in final_faces]
		tracker.step(detections)
		tracks = tracker.active_tracks(min_steps_alive=0)
		for track in tracks:
			d=[]
			d=track.box
			d_id = track.id
			d = d.astype(np.int32)
			x=d[0]
			y=d[1]
			w=d[2]
			h=d[3]
			if d_id not in id_dict.keys():
			
				#face = frame[y:h,x:w]
				
				face=frame[max(0,y-padding):min(h+padding,frame.shape[0]-1),max(0,x-padding):min(w+padding,frame.shape[1]-1)]
	
				blob=cv2.dnn.blobFromImage(face, 1.0, (227,227), MODEL_MEAN_VALUES, swapRB=False)
				genderNet.setInput(blob)
				genderPreds=genderNet.forward()
				gender=genderList[genderPreds[0].argmax()]
				print('Person :',count+1)
				print(f'Gender: {gender}')

				ageNet.setInput(blob)
				agePreds=ageNet.forward()
				age=ageList[agePreds[0].argmax()]

				print(f'Age: {age[1:-1]} years')
				id_dict[d_id] = gender
	             
				c1 = ws.cell(row = rowno, column = 1)  
				c2 = ws.cell(row = rowno, column = 2)
				c3 = ws.cell(row = rowno, column = 3)
				c4 = ws.cell(row = rowno, column = 4)

				c1.value = dateVar
				c2.value = timeVar
				c3.value = gender
				c4.value = age[1:-1]
				
				wb.save('./Output/outputgad.xlsx')

				count+=1
				rowno+=1
	                
			#	cv2.imwrite("{0}/{1}{2}_{3}.jpg".format("facepics",gender,age,d_id),face)
	
						
		#resultframe = cv2.resize(frame, (0, 0), fx=show_rate, fy=show_rate)
		for faceBox in final_faces:

			face2=frame[max(0,faceBox[1]-padding):min(faceBox[3]+padding,frame.shape[0]-1),max(0,faceBox[0]-padding):min(faceBox[2]+padding,frame.shape[1]-1)]

			blob2=cv2.dnn.blobFromImage(face2, 1.0, (227,227), MODEL_MEAN_VALUES, swapRB=False)
			genderNet.setInput(blob2)
			genderPredsshow=genderNet.forward()
			showgender=genderList[genderPredsshow[0].argmax()]

			ageNet.setInput(blob2)
			agePredsshow=ageNet.forward()
			showage=ageList[agePredsshow[0].argmax()]
			
			cv2.putText(frame, f'{showgender}, {showage}', (faceBox[0], faceBox[1]-10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0,255,255), 2, cv2.LINE_AA)
			
			cv2.rectangle(frame, (faceBox[0],faceBox[1]), (faceBox[2]-faceBox[0],faceBox[3]-faceBox[1]), (0,255,0), 2)

			cv2.imshow("Detecting age and gender",frame)	
		
		if cv2.waitKey(1) & 0XFF == ord('q'):
			print("Detection Stopped Manually")
			webcam.release()
			cv2.destroyAllWindows()
			break
	
