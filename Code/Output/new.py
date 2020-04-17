import xlrd
import datetime
from openpyxl import *
import time

#Loading the workbook

wb=load_workbook("sample.xlsx")
list=wb.get_sheet_names()

#Deleting the existing output sheet

if len(list)==2:
	std=wb.get_sheet_by_name('Sheet2')
	wb.remove_sheet(std)
	wb.save("sample.xlsx")
	
#Creating the new output sheet

wb.create_sheet("Sheet2")
ws=wb["Sheet2"]
wcell1=ws.cell(1,1)
wcell2=ws.cell(1,2)
wcell3=ws.cell(1,3)
wcell4=ws.cell(1,4)
wcell1.value="Date"
wcell2.value="Time"
wcell3.value="Group"
wcell4.value="Count"
wb.save("sample.xlsx")
book=xlrd.open_workbook("sample.xlsx")
sheet=book.sheet_by_index(0)
ws=wb["Sheet2"]

#dictionaries for finding the groups

dict1={'Male':0,'Female':8}
dict2={'0-10':1,'10-20':2,'20-30':3,'30-40':4,'40-50':5,'50-60':6,'60-70':7,'70-80':8}

for rownum1 in range(1,sheet.nrows):
	book=xlrd.open_workbook("sample.xlsx")
	sheet2=book.sheet_by_index(1)
	date=sheet.cell(rownum1,0)
	time=sheet.cell(rownum1,1)
	time_tuple = xlrd.xldate_as_tuple(time.value,book.datemode)
	gender=sheet.cell_value(rownum1, 2)
	age=sheet.cell_value(rownum1,3)
	group=dict1.get(gender)+dict2.get(age)
	flag=0
	try:
		for rownum2 in range(1,sheet2.nrows):
			date2=sheet2.cell(rownum2,0)
			time2=sheet2.cell(rownum2,1)
			time_tuple2 = xlrd.xldate_as_tuple(time2.value,book.datemode)
			group2=sheet2.cell_value(rownum2,2)
			
			#Updating the excel sheet
			
			if ((xlrd.xldate_as_tuple(date.value,book.datemode))==(xlrd.xldate_as_tuple(date2.value,book.datemode)))and time_tuple[3]==time_tuple2[3] and group==group2:
				flag=1
				count=sheet2.cell_value(rownum2,3)
				wcell=ws.cell(rownum2+1,4)
				wcell.value=count+1
				print("Updated :","Date ",date1," Time ",timec," Group ",group," Count ",count+1)
				wb.save("sample.xlsx")
				break
	except IndexError:
		print()
		
		#Creating the new row to sheet
		
	if flag==0:
		date_tuple=(xlrd.xldate_as_tuple(date.value,book.datemode))		
		date1 = datetime.date(month=date_tuple[1], day=date_tuple[2],year=date_tuple[0])
		timec=datetime.time(hour=time_tuple[3],minute=0,second=0)
		wcell1=ws.cell(sheet2.nrows+1,1)
		wcell2=ws.cell(sheet2.nrows+1,2)
		wcell3=ws.cell(sheet2.nrows+1,3)
		wcell4=ws.cell(sheet2.nrows+1,4)
		wcell1.value=date1 
		wcell2.value=timec
		wcell3.value=group
		wcell4.value=1
		print("Copied :","Date ",date1," Time ",timec," Group ",group," Count ",1)
		wb.save("sample.xlsx")
