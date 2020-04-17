import xlrd
import datetime
from openpyxl import *
import time

#Loading the workbook
sample=input("Enter the file name  ")

wb=load_workbook(sample)
	
#Creating the new output sheet

wb.create_sheet("Sheet2")
ws=wb["Sheet2"]
wcell1=ws.cell(1,1)
wcell2=ws.cell(1,2)
wcell3=ws.cell(1,3)
wcell4=ws.cell(1,4)
wcell1.value="Date"
wcell2.value="Time"
wcell3.value="Count"
wcell4.value="GroupNo"
wb.save(sample)
book=xlrd.open_workbook(sample)
sheet=book.sheet_by_index(0)
ws=wb["Sheet2"]

#dictionaries for finding the groups

dict1={'Male':0,'Female':1}
dict2={'0-10':1,'10-15':3,'15-18':5,'18-25':7,'30-45':9,'48-55':11,'60-100':13}

for rownum1 in range(1,sheet.nrows):
	book=xlrd.open_workbook(sample)
	sheet2=book.sheet_by_index(1)
	
	date=sheet.cell_value(rownum1,0)
	time=sheet.cell_value(rownum1,1)
	gender=sheet.cell_value(rownum1, 2)
	age=sheet.cell_value(rownum1,3)
	group=dict1.get(gender)+dict2.get(age)
	
	flag=0
	try:
		for rownum2 in range(1,sheet2.nrows):
			date2=sheet2.cell_value(rownum2,0)
			time2=sheet2.cell_value(rownum2,1)
			group2=sheet2.cell_value(rownum2,3)
			
			#Updating the excel sheet
			
			if date==date2 and time[0]==time2[0] and time[1]==time2[1] and group==group2:
				flag=1
				count=sheet2.cell_value(rownum2,2)
				wcell=ws.cell(rownum2+1,3)
				wcell.value=count+1
				print("Updated :","Date ",date," Time ",time[0]+time[1]," Group ",group," Count ",count+1)
				wb.save(sample)
				break
	except IndexError:
		print()
		
		#Creating the new row to sheet
		
	if flag==0:
		wcell1=ws.cell(sheet2.nrows+1,1)
		wcell2=ws.cell(sheet2.nrows+1,2)
		wcell3=ws.cell(sheet2.nrows+1,3)
		wcell4=ws.cell(sheet2.nrows+1,4)
		wcell1.value=date
		wcell2.value=time[0]+time[1]+":00:00"
		wcell3.value=1
		wcell4.value=group
		print("Copied :","Date ",date," Time ",time[0]+time[1]," Group ",group," Count ",1)
		wb.save(sample)
wb2 =load_workbook("Output.xlsx")

#removing existing sheet

sheet_name = 'Sheet1'
idx = wb2.sheetnames.index(sheet_name)
ws3 = wb2.get_sheet_by_name(sheet_name)
wb2.remove(ws3)
wb2.create_sheet(sheet_name, idx)
wb2.save("Output.xlsx")

ws2 =wb2["Sheet1"]
#copying to new sheet
for row in ws:
    for cell in row:
        ws2[cell.coordinate].value = cell.value
wb2.save("Output.xlsx")
std=wb.get_sheet_by_name('Sheet2')
wb.remove_sheet(std)
wb.save(sample)
