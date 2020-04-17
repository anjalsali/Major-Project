import cv2
import argparse
import numpy as np
import openpyxl
from openpyxl import load_workbook
from motpy import Detection, MultiObjectTracker

def run(videoName,dateVar,timeVar):

	ageProto="/root/Project Metro/Models/age_deploy.prototxt"
	ageModel="/root/Project Metro/Models/age_net.caffemodel"
	genderProto="/root/Project Metro/Models/gender_deploy.prototxt"
	genderModel="/root/Project Metro/Models/gender_net.caffemodel"
	MODEL_MEAN_VALUES=(78.4263377603, 87.7689143744, 114.895847746)
	ageList=['(0-3)', '(4-6)', '(8-15)', '(15-18)', '(18-25)', '(30-45)', '(48-55)', '(60-100)']
	genderList=['Male','Female']
	ageNet=cv2.dnn.readNet(ageModel,ageProto)
	genderNet=cv2.dnn.readNet(genderModel,genderProto)
	# initialize face detector
	face_detector = cv2.CascadeClassifier("/root/Project Metro/Models/haarcascade_frontalface_default.xml")

	detect_interval = 1
	scale_rate = 0.75
	show_rate = 1
	colours = np.random.rand(32, 3)
	tracker = MultiObjectTracker(
        dt=1/25, tracker_kwargs={'max_staleness': 3},
        model_spec='constant_acceleration_and_static_box_size_2d',
        matching_fn_kwargs={'min_iou': 0.25})

		
	c=0
	id_dict={}


	wb=openpyxl.load_workbook('outputgad.xlsx')

	ws=wb.worksheets[0]
	rowno = ws.max_row + 1
	count = ws.max_row - 1

	webcam = cv2.VideoCapture(videoName)
	fps = webcam.get(cv2.CAP_PROP_FPS)

	if fps==0:
		print("No Input Stream Detected")
		webcam.release()
		cv2.destroyAllWindows()
		return 
		
	
	tracker = MultiObjectTracker(
        dt=1/fps, tracker_kwargs={'max_staleness': 3},
        model_spec='constant_acceleration_and_static_box_size_2d',
        matching_fn_kwargs={'min_iou': 0.25})

	if not webcam.isOpened():
		print("No Input Stream Detected")
		webcam.release()
		cv2.destroyAllWindows()
		return 

	padding=20

	final_faces = []
	while(webcam.isOpened()):
		status, frame = webcam.read()
		#frame = cv2.resize(frame, (0, 0), fx=scale_rate, fy=scale_rate)
		if not status:
			print("Could not read frame")
			webcam.release()
			cv2.destroyAllWindows()
			return 
		gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
		if c % detect_interval == 0:
			#faces = face_detector.detectMultiScale(gray, 1.3, 5)
			faces = face_detector.detectMultiScale(gray,scaleFactor=2,minNeighbors=5,minSize=(300, 300))
			faces=np.array(faces)    
			face_sums = faces.shape[0]
			if face_sums > 0:
				face_list = []
				for item in faces:
					xmin = item[0]
					ymin = item[1]
					xmax = item[0] + item[2]
					ymax = item[1] + item[3]
					face_list.append([xmin, ymin, xmax, ymax])
	
				final_faces = np.array(face_list) 
		detections = [Detection(box=bbox) for bbox in final_faces]
		tracker.step(detections)
		tracks = tracker.active_tracks(min_steps_alive=0)
		for track in tracks:
			d=[]
			d=track.box
			d_id = track.id
			d = d.astype(np.int32)
			x=d[0]
			y=d[1]
			w=d[2]
			h=d[3]
			if d_id not in id_dict.keys():
				face=frame[max(0,y-padding):min(h+padding,frame.shape[0]-1),max(0,x-padding):min(w+padding, frame.shape[1]-1)]
	
				blob=cv2.dnn.blobFromImage(face, 1.0, (227,227), MODEL_MEAN_VALUES, swapRB=False)
				genderNet.setInput(blob)
				genderPreds=genderNet.forward()
				gender=genderList[genderPreds[0].argmax()]
				print('Person :',count+1)
				print(f'Gender: {gender}')

				ageNet.setInput(blob)
				agePreds=ageNet.forward()
				age=ageList[agePreds[0].argmax()]

				print(f'Age: {age[1:-1]} years')
				id_dict[d_id] = gender
	             
				c1 = ws.cell(row = rowno, column = 1)  
				c2 = ws.cell(row = rowno, column = 2)
				c3 = ws.cell(row = rowno, column = 3)
				c4 = ws.cell(row = rowno, column = 4)

				c1.value = dateVar
				c2.value = timeVar
				c3.value = gender
				c4.value = age[1:-1]
				
				wb.save('outputgad.xlsx')

				count+=1
				rowno+=1
	                
				cv2.imwrite("{0}/{1}{2}_{3}.jpg".format("facepics",gender,age,d_id),face)
	
			cv2.rectangle(frame, (x,y), (w,h), (0,255,0), 2)
	
			cv2.putText(frame, f'{gender}, {age}', (d[0], d[1]-10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,255,255), 2, cv2.LINE_AA)
			
			resultframe = cv2.resize(frame, (0, 0), fx=show_rate, fy=show_rate)
			cv2.imshow("Detecting age and gender", resultframe)
		
		if cv2.waitKey(1) & 0XFF == ord('q'):
			print("Detection Stopped Manually")
			webcam.release()
			cv2.destroyAllWindows()
			return 
			break
	
